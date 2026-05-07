#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Export Utilities

Handles data export operations:
- Summary CSV generation
- Excel chart creation
- PNG waveform export
"""

import pandas as pd
from pathlib import Path
from datetime import datetime
from typing import Optional, Dict
from openpyxl import load_workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Font, Alignment


class ExportUtils:
    """Export utilities for data and charts"""
    
    @staticmethod
    def export_summary_csv(data_path: str, output_path: str, stats: Dict) -> bool:
        """
        Export summary statistics to CSV
        
        Args:
            data_path: Path to source data file (CSV or Excel)
            output_path: Path for summary CSV output
            stats: Statistics dictionary
            
        Returns:
            bool: True if exported successfully
        """
        try:
            # Read source data
            if data_path.endswith('.csv'):
                df = pd.read_csv(data_path)
            elif data_path.endswith('.xlsx'):
                df = pd.read_excel(data_path)
            else:
                print(f"Unsupported file format: {data_path}")
                return False
            
            # Create summary data
            summary_data = {
                'Metric': [
                    'Temperature Min (°C)',
                    'Temperature Max (°C)',
                    'Temperature Avg (°C)',
                    'Humidity Min (%)',
                    'Humidity Max (%)',
                    'Humidity Avg (%)',
                    'Total Records',
                    'Sample Rate (Hz)',
                    'Session Start',
                    'Session End',
                    'Duration (minutes)'
                ],
                'Value': [
                    f"{stats.get('temp_min', 0):.1f}",
                    f"{stats.get('temp_max', 0):.1f}",
                    f"{stats.get('temp_avg', 0):.1f}",
                    f"{stats.get('humi_min', 0):.1f}",
                    f"{stats.get('humi_max', 0):.1f}",
                    f"{stats.get('humi_avg', 0):.1f}",
                    str(stats.get('packet_count', len(df))),
                    f"{stats.get('sample_rate', 0):.2f}",
                    df['Timestamp'].iloc[0] if len(df) > 0 else 'N/A',
                    df['Timestamp'].iloc[-1] if len(df) > 0 else 'N/A',
                    f"{len(df) * 2 / 60:.1f}" if len(df) > 0 else '0'  # Assuming 2-second interval
                ]
            }
            
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_csv(output_path, index=False)
            
            print(f"✓ Summary exported to: {output_path}")
            return True
            
        except Exception as e:
            print(f"Error exporting summary: {e}")
            return False
    
    @staticmethod
    def export_excel_with_chart(data_path: str, output_path: str) -> bool:
        """
        Export Excel file with embedded chart
        
        Args:
            data_path: Path to source CSV data
            output_path: Path for Excel output with chart
            
        Returns:
            bool: True if exported successfully
        """
        try:
            # Read CSV data
            df = pd.read_csv(data_path)
            
            # Create Excel with data sheet
            df.to_excel(output_path, index=False, sheet_name='Data')
            
            # Load workbook and add chart
            wb = load_workbook(output_path)
            ws = wb['Data']
            
            # Create new sheet for chart
            chart_sheet = wb.create_sheet('Chart')
            
            # Create line chart
            chart = LineChart()
            chart.title = "Temperature & Humidity Over Time"
            chart.style = 10
            chart.y_axis.title = 'Temperature (°C)'
            chart.x_axis.title = 'Sample Number'
            
            # Add data to chart
            data_rows = len(df) + 1  # +1 for header
            
            # Temperature series
            temp_data = Reference(ws, min_col=2, min_row=1, max_row=data_rows)
            chart.add_data(temp_data, titles_from_data=True)
            
            # Humidity series (on secondary axis)
            humi_data = Reference(ws, min_col=3, min_row=1, max_row=data_rows)
            chart.add_data(humi_data, titles_from_data=True)
            
            # Add chart to chart sheet
            chart_sheet.add_chart(chart, "A1")
            
            # Save workbook
            wb.save(output_path)
            wb.close()
            
            print(f"✓ Excel with chart exported to: {output_path}")
            return True
            
        except Exception as e:
            print(f"Error exporting Excel chart: {e}")
            return False
    
    @staticmethod
    def export_waveform_png(waveform_widget, output_path: str) -> bool:
        """
        Export waveform widget to PNG image
        
        Args:
            waveform_widget: WaveformWidget instance
            output_path: Path for PNG output
            
        Returns:
            bool: True if exported successfully
        """
        try:
            # Get the plot widget
            plot_widget = waveform_widget.plot_widget
            
            # Capture the rendered PlotWidget as a pixmap and save as PNG.
            pixmap = plot_widget.grab()
            if pixmap.isNull():
                print("Error exporting waveform PNG: captured image is empty")
                return False

            if not pixmap.save(output_path, "PNG"):
                print(f"Error exporting waveform PNG: failed to save to {output_path}")
                return False
            
            print(f"✓ Waveform PNG exported to: {output_path}")
            return True
            
        except Exception as e:
            print(f"Error exporting waveform PNG: {e}")
            return False
    
    @staticmethod
    def generate_export_filename(prefix: str, extension: str) -> str:
        """
        Generate export filename with timestamp
        
        Args:
            prefix: Filename prefix
            extension: File extension
            
        Returns:
            str: Generated filename
        """
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        return f"{prefix}_{timestamp}.{extension}"


if __name__ == "__main__":
    """Test export utilities"""
    import numpy as np
    
    print("=== Export Utilities Test ===\n")
    
    # Create test data
    print("1. Creating test data...")
    test_data = {
        'Timestamp': pd.date_range('2024-01-01 10:00:00', periods=50, freq='2S'),
        'Temperature': np.random.uniform(20, 30, 50),
        'Humidity': np.random.uniform(50, 70, 50),
        'OLED_State': np.random.randint(0, 2, 50)
    }
    df = pd.DataFrame(test_data)
    
    # Save test CSV
    test_csv = "./test_export_data.csv"
    df.to_csv(test_csv, index=False)
    print(f"   Test CSV created: {test_csv}")
    
    # Test 2: Export summary
    print("\n2. Exporting summary CSV...")
    test_stats = {
        'temp_min': df['Temperature'].min(),
        'temp_max': df['Temperature'].max(),
        'temp_avg': df['Temperature'].mean(),
        'humi_min': df['Humidity'].min(),
        'humi_max': df['Humidity'].max(),
        'humi_avg': df['Humidity'].mean(),
        'packet_count': len(df),
        'sample_rate': 0.5
    }
    
    summary_path = "./test_summary.csv"
    if ExportUtils.export_summary_csv(test_csv, summary_path, test_stats):
        print(f"   ✓ Summary created")
    
    # Test 3: Export Excel with chart
    print("\n3. Exporting Excel with chart...")
    excel_path = "./test_chart.xlsx"
    if ExportUtils.export_excel_with_chart(test_csv, excel_path):
        print(f"   ✓ Excel chart created")
    
    # Test 4: Generate filenames
    print("\n4. Generate export filenames:")
    filename1 = ExportUtils.generate_export_filename("summary", "csv")
    filename2 = ExportUtils.generate_export_filename("waveform", "png")
    print(f"   Summary: {filename1}")
    print(f"   Waveform: {filename2}")
    
    # Cleanup
    print("\n5. Cleanup test files...")
    for file in [test_csv, summary_path, excel_path]:
        if Path(file).exists():
            Path(file).unlink()
    print("   Test files removed")
    
    print("\n✓ Test completed!")
